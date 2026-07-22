#!/usr/bin/env python3
"""Extract Eveni templates from Excel into templates.json."""
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Instale openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

XLSX = Path("/workspace/uploads/Eveni-Templates-Projecto_v2.xlsx")
OUT = Path(__file__).resolve().parent.parent / "data" / "templates.json"

wb = openpyxl.load_workbook(XLSX, data_only=True)
print("Sheets:", wb.sheetnames)

# Identify template sheets by prefix
sheet_map = {}
for sn in wb.sheetnames:
    for tid in ("T01", "T02", "T03", "T04"):
        if sn.startswith(tid):
            sheet_map[tid] = sn

print("Detected:", sheet_map)


def parse_offset(v):
    if v is None:
        return None
    s = str(v).strip().upper()
    if not s:
        return None
    # Accept "D-90", "D+7", "D0", "-90", "90"
    if s.startswith("D"):
        s = s[1:]
    if s == "" or s == "0":
        return 0
    try:
        return int(s)
    except ValueError:
        return None


templates = {}
for tid, sname in sheet_map.items():
    ws = wb[sname]
    tasks = []
    header_row = None
    # Find header row: first row where first cell equals "ID" (any case)
    for row_idx in range(1, min(10, ws.max_row + 1)):
        first = ws.cell(row=row_idx, column=1).value
        if first and str(first).strip().lower() == "id":
            header_row = row_idx
            break
    if header_row is None:
        header_row = 1

    # Detect columns
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[header_row]]
    def col(*names):
        for n in names:
            for i, h in enumerate(headers):
                if n in h:
                    return i
        return None

    ic_id = col("id")
    ic_phase = col("fase")
    ic_cat = col("categoria")
    ic_title = col("título", "titulo")
    ic_role = col("papel", "responsável", "responsavel")
    ic_start = col("início", "inicio", "start")
    ic_end = col("fim", "end")
    ic_crit = col("crítica", "critica", "critical")
    ic_note = col("nota", "obs", "note")

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        title = row[ic_title] if ic_title is not None and ic_title < len(row) else None
        if not title:
            continue
        title_s = str(title).strip()
        if not title_s:
            continue
        task = {
            "id": str(row[ic_id]).strip() if ic_id is not None and row[ic_id] else None,
            "phase": str(row[ic_phase]).strip() if ic_phase is not None and row[ic_phase] else "",
            "category": str(row[ic_cat]).strip() if ic_cat is not None and row[ic_cat] else "",
            "title": title_s,
            "role": str(row[ic_role]).strip() if ic_role is not None and row[ic_role] else None,
            "offset_start": parse_offset(row[ic_start]) if ic_start is not None else None,
            "offset_end": parse_offset(row[ic_end]) if ic_end is not None else None,
            "critical": bool(row[ic_crit]) if ic_crit is not None and row[ic_crit] else False,
            "note": str(row[ic_note]).strip() if ic_note is not None and row[ic_note] else "",
        }
        tasks.append(task)

    templates[tid] = {
        "name": {
            "T01": "Conferência de Imprensa",
            "T02": "Casamento",
            "T03": "Congresso",
            "T04": "Festival / Concerto",
        }.get(tid, tid),
        "sheet": sname,
        "tasks": tasks,
    }
    print(f"{tid} ({sname}): {len(tasks)} tarefas")

OUT.parent.mkdir(parents=True, exist_ok=True)
OUT.write_text(json.dumps(templates, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"Wrote {OUT}")
